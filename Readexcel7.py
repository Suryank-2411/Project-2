import openpyxl as xl 
import os
def get_no_of_rows(worksheet):
    count=0
    for k in range(1,(worksheet.max_row)+1):
        if(worksheet.cell(row=k,column=1).value == None):
            count=count+1
            
    no_of_rows = (worksheet.max_row) - count
    return no_of_rows
def get_no_of_columns(worksheet):
    countc=0
    for l in range(1,(worksheet.max_column)+1):
        if(worksheet.cell(row=1,column=l).value == None):
            countc=countc+1
    no_of_cols = (worksheet.max_column) - countc
    return no_of_cols
def temporunner(i,cws):                  # to clean the excel sheet
    for k in range((i%1000000)+1,1000001):
        print("going on at k="+str(k))
        for l in range(1,maximum_columns+1):
            print("Moving on with l="+str(l))
            cws.cell(row=k,column=l).value = None
path="/mnt/c/Users/surya_h3yma/Autodesk/Python/Excelsheets/.xlsx/.xlsx" 
os.chdir(path)
# opening the workbook in which data has to be copied
# cwb= copy workbook, cws= copy worksheet
cwb = xl.load_workbook("/mnt/c/Users/surya_h3yma/Autodesk/Python"+"/"+"copywb.xlsx")
cws= cwb.active
list_of_workbooks = os.listdir(path)
print("list of workbooks")
print(list_of_workbooks)
for file in list_of_workbooks:
    print(file)
    wb = xl.load_workbook(path +"/"+file)
    sheet_names = wb.sheetnames       #list of all the sheets present in the workbook,with sheet names in it
    no_of_worksheets = len(sheet_names)
    print("no. of worksheets",end="=")
    print(no_of_worksheets)
    for index in range(no_of_worksheets):
        ws = wb.worksheets[index]
        print(ws) # got the worksheet
        maximum_rows = get_no_of_rows(ws)
        print("maximum rows ="+str(maximum_rows))
        maximum_columns= get_no_of_columns(ws)
        print("maximum columns="+str(maximum_columns))
        i=0
        j=0
        loop_value_rows=0
        for i in range(1,maximum_rows+1):
            print("no of preccesing")
            print(i)
            for j in range(1,maximum_columns+1):
                header_value =str(ws.cell(row=1,column=j).value) #  copying the value of head elements of all the rows to check whether name or number is present
                print(header_value)
                if "name" in str(header_value):
                    c=ws.cell(row=i,column=j)
                    cws.cell(row=i-loop_value_rows,column=j).value = c.value
                elif "Name" in str(header_value):
                    c=ws.cell(row=i,column=j)
                    cws.cell(row=i-loop_value_rows,column=j).value = c.value
                elif "NAME" in str(header_value):
                    c=ws.cell(row=i,column=j)
                    cws.cell(row=i-loop_value_rows,column=j).value = c.value
                elif "NUMBER" in str(header_value):
                    c=ws.cell(row=i,column=j)
                    cws.cell(row=i-loop_value_rows,column=j).value = c.value
                elif "NO" in str(header_value):
                    c=ws.cell(row=i,column=j)
                    cws.cell(row=i-loop_value_rows,column=j).value = c.value
                elif "No" in str(header_value):
                    c=ws.cell(row=i,column=j)
                    cws.cell(row=i-loop_value_rows,column=j).value = c.value
                elif "Number" in str(header_value):
                    c=ws.cell(row=i,column=j)
                    cws.cell(row=i-loop_value_rows,column=j).value = c.value
                elif "number" in str(header_value):
                    c=ws.cell(row=i,column=j)
                    cws.cell(row=i-loop_value_rows,column=j).value = c.value
                elif "MOBILE" in str(header_value):
                    c=ws.cell(row=i,column=j)
                    cws.cell(row=i-loop_value_rows,column=j).value = c.value
                elif "mobile" in str(header_value):
                    c=ws.cell(row=i,column=j)
                    cws.cell(row=i-loop_value_rows,column=j).value = c.value
                elif "Mobile" in str(header_value):
                    c=ws.cell(row=i,column=j)
                    cws.cell(row=i-loop_value_rows,column=j).value = c.value
                else:
                    cws.cell(row=i-loop_value_rows,column=j).value = ""
            if((i)%1000000==0):
                cwb.save("copywb"+"_"+str(i)+"_"+"("+file+")"+"("+str(sheet_names[index])+")"+".xlsx")
                loop_value_rows=i
            elif(i%1000000!=0 and i==maximum_rows):
                cwb.save("copywb"+"_"+"final"+"_"+"("+file+")"+"("+str(sheet_names[index])+")"+".xlsx")
                temporunner(i,cws)                
#def temporunner(i,cws):
#    for k in range((i%1000000)+1,1000001):
#        print("going on at k="+str(k))
#        for l in range(1,maximum_columns+1):
#            print("Moving on with l="+str(l))
#            cws.cell(row=k,column=l).value = None
try:
    header_value = None
    if "name" in header_value:
        print("Error handled")
except Exception :
    print("Skipped this column")





        
